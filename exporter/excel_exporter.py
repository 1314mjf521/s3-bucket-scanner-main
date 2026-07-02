"""Excel导出器实现"""
import os
from datetime import datetime
from typing import List, Optional
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from model.s3_object import S3Object
from exporter.exporter import ExcelExporter


class ExcelFileHandler:
    """Excel文件处理器，支持大文件分割"""
    
    def __init__(
        self,
        base_filename: str,
        max_rows_per_sheet: int = 1000000,
        max_columns: int = 26
    ):
        self.base_filename = base_filename
        self.max_rows_per_sheet = max_rows_per_sheet
        self.max_columns = max_columns
        self.current_workbook: Optional[Workbook] = None
        self.current_sheet = None
        self.current_row = 0
        self.file_index = 1
        self.total_rows_written = 0
        
    def _create_new_workbook(self) -> None:
        """创建新的工作簿"""
        self.current_workbook = Workbook()
        # 使用 filter_sheet_name 清理工作表名称
        from util.filename_sanitizer import filter_sheet_name
        default_name = self.current_workbook.active.title
        sanitized_name = filter_sheet_name(default_name)
        self.current_workbook.active.title = sanitized_name
        self.current_sheet = self.current_workbook.active
        self.current_row = 0
        self._setup_sheet_headers()
        # After _setup_sheet_headers, current_row is 1 (header row)
        # We want to start writing from row 2
        self.current_row = 0  # Reset to 0 for proper counting
        
    def _setup_sheet_headers(self) -> None:
        """设置表头"""
        headers = ['Key', 'Bucket', 'Path', 'Last Modified', 'Size (bytes)', 'ETag']
        for col, header in enumerate(headers, 1):
            self.current_sheet.cell(row=1, column=col, value=header)
        self.current_row = 1
        
    def _get_filename(self) -> str:
        """获取当前文件名"""
        # 清理文件名中的非法字符
        sanitized_base = filter_filename(self.base_filename)
        
        if self.file_index == 1:
            return f"{sanitized_base}.xlsx"
        return f"{sanitized_base}_{self.file_index}.xlsx"
        
    def _check_and_split(self) -> bool:
        """检查是否需要分割文件"""
        if self.current_row >= self.max_rows_per_sheet:
            self.file_index += 1
            self._create_new_workbook()
            return True
        return False
        
    def write_row(self, data: dict) -> None:
        """写入一行数据"""
        if self.current_sheet is None:
            self._create_new_workbook()
            
        self._check_and_split()
        self.current_row += 1
        self.total_rows_written += 1
        
        # 直接写入原始数据，不进行过滤
        # openpyxl 支持 # 和 & 字符，只过滤控制字符
        key = data.get('key', '')
        path = data.get('path', '')
        
        # 写入数据，如果失败则记录日志并跳过
        try:
            self.current_sheet.cell(row=self.current_row + 1, column=1, value=key)
            self.current_sheet.cell(row=self.current_row + 1, column=2, value=data.get('bucket', ''))
            self.current_sheet.cell(row=self.current_row + 1, column=3, value=path)
            self.current_sheet.cell(row=self.current_row + 1, column=4, value=str(data.get('last_modified', '')))
            self.current_sheet.cell(row=self.current_row + 1, column=5, value=data.get('size', 0))
            self.current_sheet.cell(row=self.current_row + 1, column=6, value=data.get('etag', ''))
        except Exception as e:
            # 记录无法处理的行并跳过
            print(f"Warning: Skipping row with problematic key: {repr(key)}")
            print(f"Warning: Error: {e}")
            self.current_row -= 1
            self.total_rows_written -= 1
        
    def save(self) -> List[str]:
        """保存所有工作簿，返回文件列表"""
        files_saved = []
        
        if self.current_workbook is not None:
            filename = self._get_filename()
            try:
                self.current_workbook.save(filename)
                files_saved.append(filename)
            except Exception as e:
                print(f"DEBUG: Error saving workbook: {e}")
                print(f"DEBUG: Filename: {filename}")
                raise
                
        return files_saved


class ExcelExporterImpl(ExcelExporter):
    """Excel导出器实现"""
    
    def __init__(
        self,
        base_filename: str = 's3_objects_export',
        max_rows_per_sheet: int = 1000000,
        batch_size: int = 1000
    ):
        self.base_filename = base_filename
        self.max_rows_per_sheet = max_rows_per_sheet
        self.batch_size = batch_size
        self.file_handler: Optional[ExcelFileHandler] = None
        self.is_open = False
        
    async def open(self) -> None:
        """打开导出器"""
        self.file_handler = ExcelFileHandler(
            base_filename=self.base_filename,
            max_rows_per_sheet=self.max_rows_per_sheet
        )
        self.is_open = True
        
    async def write(self, data: List[S3Object]) -> None:
        """写入数据（批量）"""
        if not self.is_open or self.file_handler is None:
            raise RuntimeError("Exporter is not open. Call open() first.")
            
        if not data:
            return
            
        # 批量写入
        for obj in data:
            # 调试：打印原始 key
            print(f"DEBUG: Writing object key: {repr(obj.key)}")
            self.file_handler.write_row({
                'key': obj.key,
                'bucket': obj.bucket,
                'path': obj.path,
                'last_modified': obj.last_modified,
                'size': obj.size,
                'etag': obj.etag,
                'debug_key_check': True  # 启用调试
            })
            
    async def close(self) -> List[str]:
        """关闭导出器并保存文件"""
        if not self.is_open:
            return []
            
        files_saved = self.file_handler.save()
        self.is_open = False
        self.file_handler = None
        
        return files_saved
        
    def set_file_name(self, file_name: str) -> None:
        """设置文件名"""
        self.base_filename = file_name
        
    def set_max_rows_per_sheet(self, max_rows: int) -> None:
        """设置每张表最大行数"""
        self.max_rows_per_sheet = max_rows
